"""
Hierarchical document chunking service.
Implements hierarchical chunking approach for document segmentation.
"""
from typing import List, Dict, Any
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
from app.core.config import settings
from app.core.logging import get_logger

logger = get_logger(__name__)


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file."""
    try:
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {e}")
        raise


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX file."""
    try:
        doc = DocxDocument(file_path)
        text = "\n\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])
        return text
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        raise


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from XLSX file."""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text
    except Exception as e:
        logger.error(f"Error extracting text from XLSX: {e}")
        raise


def extract_text_from_txt(file_path: str) -> str:
    """Extract text from TXT file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except UnicodeDecodeError:
        # Try with different encoding if UTF-8 fails
        try:
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error extracting text from TXT: {e}")
            raise
    except Exception as e:
        logger.error(f"Error extracting text from TXT: {e}")
        raise


def extract_text_from_document(file_path: str, file_type: str) -> str:
    """
    Extract text from document based on file type.

    Args:
        file_path: Path to the document file
        file_type: MIME type of the file

    Returns:
        Extracted text content
    """
    if file_type == "application/pdf":
        return extract_text_from_pdf(file_path)
    elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return extract_text_from_docx(file_path)
    elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return extract_text_from_xlsx(file_path)
    elif file_type == "text/plain":
        return extract_text_from_txt(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def hierarchical_chunking(
    document_text: str,
    chunk_size: int = None,
    chunk_overlap: int = None
) -> List[Dict[str, Any]]:
    """
    Implement hierarchical chunking approach.
    Splits document into overlapping chunks with context.

    Args:
        document_text: The full document text
        chunk_size: Maximum size of each chunk (default from settings)
        chunk_overlap: Overlap between chunks (default from settings)

    Returns:
        List of chunk dictionaries with metadata
    """
    if chunk_size is None:
        chunk_size = settings.chunk_size
    if chunk_overlap is None:
        chunk_overlap = settings.chunk_overlap

    chunks = []

    # Split by paragraphs first
    paragraphs = [p.strip() for p in document_text.split('\n\n') if p.strip()]

    current_chunk = ""
    chunk_index = 0

    for para_idx, paragraph in enumerate(paragraphs):
        # If adding this paragraph would exceed chunk size
        if len(current_chunk) + len(paragraph) > chunk_size and current_chunk:
            # Save current chunk
            chunks.append({
                "content": current_chunk.strip(),
                "chunk_index": chunk_index,
                "start_para": para_idx - len(current_chunk.split('\n\n')),
                "end_para": para_idx,
                "char_count": len(current_chunk)
            })

            # Start new chunk with overlap
            # Take last N characters for overlap
            overlap_text = current_chunk[-chunk_overlap:] if len(current_chunk) > chunk_overlap else current_chunk
            current_chunk = overlap_text + "\n\n" + paragraph
            chunk_index += 1
        else:
            # Add paragraph to current chunk
            if current_chunk:
                current_chunk += "\n\n" + paragraph
            else:
                current_chunk = paragraph

    # Add final chunk
    if current_chunk.strip():
        chunks.append({
            "content": current_chunk.strip(),
            "chunk_index": chunk_index,
            "start_para": len(paragraphs) - len(current_chunk.split('\n\n')),
            "end_para": len(paragraphs),
            "char_count": len(current_chunk)
        })

    logger.info(f"Created {len(chunks)} chunks from document")
    return chunks


def process_document(file_path: str, file_type: str) -> List[Dict[str, Any]]:
    """
    Complete document processing pipeline.
    Extracts text and creates chunks.

    Args:
        file_path: Path to document file
        file_type: MIME type of the file

    Returns:
        List of processed chunks with metadata
    """
    logger.info(f"Processing document: {file_path}")

    # Extract text
    text = extract_text_from_document(file_path, file_type)
    logger.info(f"Extracted {len(text)} characters from document")

    # Create chunks
    chunks = hierarchical_chunking(text)

    return chunks
